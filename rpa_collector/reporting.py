from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


LOGGER = logging.getLogger(__name__)


def _clean_id(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip()


def _latest_file(directory: Path, pattern: str) -> Path:
    files = list(directory.glob(pattern))
    if not files:
        raise FileNotFoundError(f"在 {directory} 中找不到文件：{pattern}")
    return max(files, key=lambda path: path.stat().st_mtime)


def _load_dictionary_ids(
    workbook: Path,
    id_column: str,
    sheet_names: list[str],
    date_column: str,
    start_date: str,
    target_column: str,
    target_value: str,
) -> set[str]:
    ids: set[str] = set()
    excel = pd.ExcelFile(workbook)
    missing_sheets = set(sheet_names) - set(excel.sheet_names)
    if missing_sheets:
        raise ValueError(f"字典工作簿缺少工作表：{sorted(missing_sheets)}")
    for sheet_name in sheet_names:
        frame = pd.read_excel(workbook, sheet_name=sheet_name, dtype=str)
        required = {id_column, date_column, target_column}
        missing = required - set(frame.columns)
        if missing:
            raise ValueError(f"字典工作表“{sheet_name}”缺少字段：{sorted(missing)}")
        dates = pd.to_datetime(frame[date_column], errors="coerce")
        targets = frame[target_column].fillna("").astype(str).str.strip()
        mask = (dates >= pd.Timestamp(start_date)) & (targets == target_value)
        ids.update(value for value in _clean_id(frame.loc[mask, id_column]) if value)
    if not ids:
        raise ValueError(f"字典工作簿中未找到有效的“{id_column}”")
    return ids


def _calculate_july_gmv(
    csv_path: Path,
    dictionary_ids: set[str],
    content_id_column: str,
    gmv_column: str,
) -> tuple[float, pd.DataFrame]:
    frame = pd.read_csv(csv_path, dtype=str)
    missing = {content_id_column, gmv_column} - set(frame.columns)
    if missing:
        raise ValueError(f"七月报表缺少字段：{sorted(missing)}")

    frame[content_id_column] = _clean_id(frame[content_id_column])
    frame[gmv_column] = pd.to_numeric(
        frame[gmv_column]
        .fillna("0")
        .astype(str)
        .str.strip()
        .str.replace(",", "", regex=False),
        errors="coerce",
    ).fillna(0.0)
    matched = frame.loc[frame[content_id_column].isin(dictionary_ids)].copy()
    return round(float(matched[gmv_column].sum()), 2), matched


def generate_gmv_report(root: Path, config: dict[str, Any]) -> Path:
    report_config = config.get("reporting", {})
    runtime_file = root / report_config.get("gmv_data_file", "runtime/extracted_data.json")
    dictionary_file = root / report_config.get(
        "dictionary_file", "Gotyasatch戈撒驰书包日报0902.xlsx"
    )
    july_csv = _latest_file(
        root / report_config.get("july_report_dir", "output/july_reports"), "*.csv"
    )

    records = json.loads(runtime_file.read_text(encoding="utf-8"))
    if not isinstance(records, list):
        raise ValueError(f"GMV 临时数据格式错误：{runtime_file}")

    id_column = str(report_config.get("dictionary_id_column", "笔记/素材ID"))
    dictionary_sheets = [
        str(name) for name in report_config.get("dictionary_sheets", ["CPUV底表"])
    ]
    content_id_column = str(report_config.get("content_id_column", "内容ID"))
    gmv_column = str(report_config.get("gmv_column", "商家GMV"))
    start_date = str(report_config.get("start_date", "2026-08-09"))
    dictionary_ids = _load_dictionary_ids(
        dictionary_file,
        id_column,
        dictionary_sheets,
        str(report_config.get("dictionary_date_column", "时间")),
        start_date,
        str(report_config.get("dictionary_target_column", "推广目标")),
        str(report_config.get("dictionary_target_value", "CPUV")),
    )
    july_gmv, matched = _calculate_july_gmv(
        july_csv, dictionary_ids, content_id_column, gmv_column
    )

    values = {str(record["project_name"]): record.get("merchant_gmv", "na") for record in records}
    values[str(report_config.get("july_project_name", "戈撒驰7月"))] = july_gmv
    project_order = report_config.get("project_order", list(values))
    rows = []
    for project_name in project_order:
        value = values.get(str(project_name), "na")
        output_value = None if value in (None, "", "na", "null") else float(value)
        rows.append({"项目名": str(project_name), "商家GMV": output_value})

    current_day = datetime.now().astimezone().date().isoformat()
    output_dir = root / report_config.get("output_dir", "output/final_reports")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"戈撒驰星河平台投流商家GMV报表_{start_date}_{current_day}.xlsx"
    if output_path.exists():
        try:
            with output_path.open("a+b"):
                pass
        except PermissionError:
            suffix = datetime.now().strftime("%H%M%S")
            output_path = output_dir / (
                f"戈撒驰星河平台投流商家GMV报表_{start_date}_{current_day}_{suffix}.xlsx"
            )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, sheet_name="GMV报表", index=False, startrow=2)
        matched.to_excel(writer, sheet_name="7月匹配明细", index=False)
        workbook = writer.book
        sheet = writer.sheets["GMV报表"]
        sheet.merge_cells("A1:B1")
        sheet["A1"] = "戈撒驰星河平台投流商家GMV报表"
        sheet.merge_cells("A2:B2")
        sheet["A2"] = f"时间维度：{start_date} to {current_day}"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet["A2"].alignment = Alignment(horizontal="center")
        for cell in sheet[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions["A"].width = 42
        sheet.column_dimensions["B"].width = 18
        for row in range(4, sheet.max_row + 1):
            sheet.cell(row=row, column=2).number_format = '#,##0.00'
        sheet.freeze_panes = "A4"
        detail_sheet = writer.sheets["7月匹配明细"]
        detail_sheet.freeze_panes = "A2"
        detail_sheet.auto_filter.ref = detail_sheet.dimensions
        workbook.calculation.fullCalcOnLoad = True

    LOGGER.info(
        "最终报表已生成：%s（七月匹配 %s 行、%s 个内容ID，GMV %.2f）",
        output_path,
        len(matched),
        matched[content_id_column].nunique(),
        july_gmv,
    )
    return output_path
